# -*- coding: utf-8 -*-
"""
교육 자동화 ① 준비·검증.

make_template.py 로 만든 '교육_입력양식.xlsx'(교육정보+대상자 시트)를 읽어서:
  - 교육정보의 필수 항목이 다 채워졌는지 확인하고
  - 대상자별로 이수=Y / 수료증 파일 매칭을 점검해 '작업목록'을 만들고
  - 수료증 누락 / 사번 누락 / 중복 / 미이수를 미리 잡아준다.
ERP 는 건드리지 않는다(안전). 이 결과를 ②RPA 단계가 그대로 사용한다.

사용법:
  python prepare.py                              (output/교육_입력양식.xlsx 자동 사용)
  python prepare.py --input 내가채운양식.xlsx
  python prepare.py --selftest                   (파일 없이 동작 확인)

수료증 폴더: training/수료증/  (파일명을 '사번.pdf' 형태로)
"""

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config

ROOT = Path(__file__).resolve().parent
CERT_DIR = ROOT / config.CERT_DIR
OUTPUT_DIR = ROOT / "output"
DEFAULT_INPUT = OUTPUT_DIR / "교육_입력양식.xlsx"

HEADER_ALIASES = {
    "사번": "사번", "사원번호": "사번", "사원": "사번",
    "성명": "성명", "이름": "성명",
    "이수여부": "이수여부", "이수": "이수여부", "수료여부": "이수여부",
    "수료증파일": "수료증", "수료증": "수료증", "파일": "수료증",
}


def read_template(path: Path):
    """입력양식(.xlsx)에서 (교육정보 dict, 대상자 list) 를 읽는다."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if "교육정보" not in wb.sheetnames or "대상자" not in wb.sheetnames:
        raise ValueError("'교육정보'/'대상자' 시트가 없습니다. make_template.py 로 만든 양식인지 확인하세요.")
    course = _first_data_row(wb["교육정보"])
    attendees = [_normalize_keys(r) for r in _all_data_rows(wb["대상자"])]
    return course, attendees


def _first_data_row(ws) -> dict:
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    for values in rows:
        if values and any(v not in (None, "") for v in values):
            return {h: ("" if v is None else str(v).strip()) for h, v in zip(headers, values)}
    return {}


def _all_data_rows(ws) -> list:
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    out = []
    for values in rows:
        if not values or all(v in (None, "") for v in values):
            continue
        out.append(dict(zip(headers, values)))
    return out


def _normalize_keys(row: dict) -> dict:
    norm = {"사번": "", "성명": "", "이수여부": "", "수료증": ""}
    for key, value in row.items():
        canon = HEADER_ALIASES.get(str(key).strip())
        if canon:
            norm[canon] = "" if value is None else str(value).strip()
    return norm


def scan_certs(cert_dir: Path) -> dict:
    """수료증 폴더를 훑어 {키: 경로} 인덱스 생성. 키 = 파일명(확장자 제외)과 '_' 앞부분."""
    index = {}
    if not cert_dir.exists():
        return index
    for path in cert_dir.iterdir():
        if path.is_file() and path.suffix.lower() in config.CERT_EXTS:
            stem = path.stem.strip()
            index.setdefault(stem, path)
            index.setdefault(stem.split("_")[0].strip(), path)  # 82211493_김동은.pdf 대응
    return index


def match_cert(row: dict, certs: dict):
    """행에 맞는 수료증 파일을 찾는다. 명시 파일명 > 사번 > 성명 순."""
    explicit = row.get("수료증", "")
    if explicit:
        direct = CERT_DIR / explicit
        if direct.exists():
            return direct
        if Path(explicit).stem in certs:
            return certs[Path(explicit).stem]
    for key in (row.get("사번", ""), row.get("성명", "")):
        if key and key in certs:
            return certs[key]
    return None


def is_done(value: str) -> bool:
    return value.strip().lower() in config.DONE_VALUES


def check_course(course: dict) -> list:
    return [field for field in config.COURSE_REQUIRED if not str(course.get(field, "")).strip()]


def build_worklist(rows: list, certs: dict) -> list:
    out, seen = [], set()
    for row in rows:
        sabun = row.get("사번", "")
        name = row.get("성명", "")
        done = is_done(row.get("이수여부", ""))
        cert = match_cert(row, certs)

        notes = []
        if not sabun:
            notes.append("사번 없음")
        if not name:
            notes.append("성명 없음")
        if sabun and sabun in seen:
            notes.append("사번 중복")
        if sabun:
            seen.add(sabun)

        if not done:
            notes.append("미이수(N)-처리 제외")
        elif config.CERT_REQUIRED and cert is None:
            notes.append("수료증 없음")

        out.append({
            "사번": sabun,
            "성명": name,
            "이수여부": "Y" if done else "N",
            "수료증파일": cert.name if cert else "",
            "상태": "; ".join(notes) if notes else "처리대상",
            "_문제": bool(notes) and done,
        })
    return out


def write_xlsx(rows: list, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "작업목록"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(config.COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in rows:
        ws.append([row.get(col, "") for col in config.COLUMNS])
        if row.get("_문제"):
            for cell in ws[ws.max_row]:
                cell.fill = warn_fill

    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def summarize(rows: list) -> None:
    targets = [r for r in rows if r["이수여부"] == "Y"]
    problems = [r for r in rows if r["_문제"]]
    print(f"  대상자 총 {len(rows)}명 / 이수대상 {len(targets)}명 / 검토필요 {len(problems)}명")
    for r in problems:
        print(f"    ! {r['사번']} {r['성명']}: {r['상태']}")


def run(input_path: Path, out_path: Path) -> None:
    if not input_path.exists():
        print(f"[안내] 입력양식이 없습니다: {input_path}")
        print("       먼저  python make_template.py  로 양식을 만들고 채운 뒤 실행하세요.")
        return

    course, attendees = read_template(input_path)
    print(f"[교육정보] {course.get('과정명') or '(과정명 비어있음)'}")
    missing = check_course(course)
    if missing:
        print("  ! 필수 항목 비어있음:", ", ".join(missing))
    else:
        print("  필수 항목 모두 입력됨 ✔")

    if not attendees:
        print("[대상자] '대상자' 시트에 사람이 없습니다.")
        return

    rows = build_worklist(attendees, scan_certs(CERT_DIR))
    write_xlsx(rows, out_path)
    print(f"[대상자] 작업목록 생성: {out_path}")
    summarize(rows)
    print("\n검토 후 문제(주황색)가 없으면 다음 단계(ERP 자동입력)로 넘어갑니다.")


def selftest(out_path: Path) -> None:
    """파일 없이 매칭/검증 로직만 점검 (수료증 인덱스는 가짜로 주입)."""
    rows = [
        {"사번": "82211493", "성명": "김동은", "이수여부": "Y", "수료증": ""},
        {"사번": "10001234", "성명": "홍길동", "이수여부": "예", "수료증": ""},  # 수료증 없음
        {"사번": "82211493", "성명": "김동은", "이수여부": "Y", "수료증": ""},  # 중복
        {"사번": "", "성명": "무사번", "이수여부": "Y", "수료증": ""},          # 사번 없음
        {"사번": "10005678", "성명": "이불참", "이수여부": "N", "수료증": ""},  # 미이수 제외
    ]
    fake_certs = {"82211493": Path("수료증/82211493.pdf")}
    worklist = build_worklist(rows, fake_certs)
    write_xlsx(worklist, out_path)
    print(f"[selftest] 샘플 5명으로 생성 완료: {out_path}")
    summarize(worklist)


def main() -> None:
    parser = argparse.ArgumentParser(description="교육 입력양식 검증 → 작업목록")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="입력양식 엑셀 경로")
    parser.add_argument("--output", default=str(OUTPUT_DIR / "작업목록.xlsx"), help="결과 경로")
    parser.add_argument("--selftest", action="store_true", help="파일 없이 동작 확인")
    args = parser.parse_args()

    out_path = Path(args.output)
    if args.selftest:
        selftest(out_path)
    else:
        run(Path(args.input), out_path)


if __name__ == "__main__":
    main()
