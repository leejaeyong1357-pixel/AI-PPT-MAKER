# -*- coding: utf-8 -*-
"""
교육 입력 양식(엑셀) 생성기.

봇이 읽을 '입력 엑셀'을 만든다. 시트 2개:
  - '교육정보' : 한 줄 = 교육 1개. (1·2단계 폼에 들어갈 항목들)
  - '대상자'   : 한 줄 = 사람 1명. (3단계 이수/증빙 처리할 사람들)
사용법:  python make_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "output" / "교육_입력양식.xlsx"

# (헤더, 예시값, 필수여부)
COURSE_FIELDS = [
    ("과정명", "M2,M1 주니어관리자 실무", True),
    ("사내외구분", "사내교육", True),          # 사내교육 / 사외교육
    ("교육구분", "기본교육", True),
    ("교육형태", "집체교육", False),
    ("교육목적", "실무 역량 강화 교육", True),
    ("교육내용", "실무 역량 강화 교육", True),
    ("시작일", "2026-02-26", True),
    ("종료일", "2026-12-31", True),
    ("교육장소", "HMG 경주연수원", True),
    ("교육시간", "16", True),
    ("교육비", "0", True),
    ("교육기관", "", False),
    ("비고", "", False),
]

ATTENDEE_FIELDS = [
    ("사번", "82211493"),
    ("성명", "김동은"),
    ("이수여부", "Y"),            # Y / N
    ("수료증파일", "82211493.pdf"),  # 수료증/ 폴더의 파일명
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
REQ_FILL = PatternFill("solid", fgColor="C6E0B4")   # 필수 칸은 연한 초록
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, headers, required=None):
    required = required or set()
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL if name not in required else REQ_FILL
        cell.alignment = CENTER
        ws.column_dimensions[cell.column_letter].width = max(len(name) + 4, 12)
    ws.freeze_panes = "A2"


def build() -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "교육정보"
    headers1 = [h for h, _, _ in COURSE_FIELDS]
    required1 = {h for h, _, req in COURSE_FIELDS if req}
    _style_header(ws1, headers1, required1)
    ws1.append([ex for _, ex, _ in COURSE_FIELDS])  # 예시 한 줄
    # 사내외구분 드롭다운
    dv = DataValidation(type="list", formula1='"사내교육,사외교육"', allow_blank=True)
    ws1.add_data_validation(dv)
    dv.add(f"B2:B200")

    ws2 = wb.create_sheet("대상자")
    headers2 = [h for h, _ in ATTENDEE_FIELDS]
    _style_header(ws2, headers2)
    ws2.append([ex for _, ex in ATTENDEE_FIELDS])
    dv2 = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws2.add_data_validation(dv2)
    dv2.add("C2:C500")

    ws3 = wb.create_sheet("작성안내")
    notes = [
        "[작성 방법]",
        "1) '교육정보' 시트: 만들 교육 1개당 한 줄. 초록색 헤더 = 필수 입력.",
        "   - 사내외구분: 사내교육 / 사외교육 중 선택",
        "   - 날짜: YYYY-MM-DD 형식 (예: 2026-02-26)",
        "2) '대상자' 시트: 이 교육을 들은 사람을 한 줄씩. (= 봇이 누를 사람들)",
        "   - 이수여부: Y(이수) / N(미이수, 처리 제외)",
        "   - 수료증파일: training/수료증/ 폴더에 같은 이름으로 두기 (예: 82211493.pdf)",
        "3) 예시로 들어있는 첫 줄은 지우거나 덮어쓰세요.",
        "",
        "[봇이 하는 일]  교육과정등록 → 교육과정개설등록 → 교육결과등록(한 명씩 Y+증빙) 자동 입력.",
        "저장 직전마다 확인을 거치며, ERP 로그인은 본인이 직접 합니다.",
    ]
    for i, line in enumerate(notes, start=1):
        ws3.cell(row=i, column=1, value=line)
    ws3.column_dimensions["A"].width = 90

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"생성 완료: {OUT}")


if __name__ == "__main__":
    build()
