# -*- coding: utf-8 -*-
"""
외국어 성적표(이미지/PDF) -> ERP 외국어 업로드용 엑셀 자동 변환기.

흐름:
  1) input/ 폴더에 성적표 사진/PDF를 넣는다.
  2) python extract.py 실행 -> AI가 각 파일을 읽어 점수/등급 등을 뽑는다.
  3) (선택) roster.csv(성명,사번)가 있으면 사번을 자동으로 채운다.
  4) 회사 규칙(config.py)으로 외국어 능력(상/중/하)을 매긴다.
  5) output/외국어_업로드.xlsx 생성 + 만료/중복/누락을 '검증결과' 칸에 표시.
  6) 사람이 한 번 검토 -> ERP 외국어 화면에 그대로 대량 업로드.

ERP 접속은 전혀 하지 않습니다. (사내망 ERP 로그인/업로드는 본인 PC에서 수동으로)
"""

import argparse
import base64
import csv
import datetime
import os
import sys
from pathlib import Path
from typing import Optional

from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config

ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "input"
OUTPUT_DIR = ROOT / "output"
IMAGE_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".gif": "image/gif",
}

SYSTEM_PROMPT = """당신은 한국 회사의 외국어 성적표(이미지 또는 PDF)에서 정보를 추출하는 도우미입니다.
주어진 증빙을 읽어 아래 항목을 구조화하세요. 보이지 않거나 확신이 없으면 반드시 null로 두세요. 절대 추측하지 마세요.

- employee_name: 응시자 이름(한글). 없으면 null
- employee_id: 사번. 성적표에는 보통 없으므로 없으면 null
- language: 외국어 구분 (예: 영어, 중국어, 일본어)
- test_type: 시험 종류를 표준 명칭으로 통일 (예: OPIc, TOEIC, TOEIC Speaking, TEPS, SPA, JPT, HSK)
- valid_date: 유효일자(만료일)를 우선. 없으면 시험일/발급일. 반드시 YYYY-MM-DD 형식
- listening / reading / writing / speaking: 영역별 점수(정수). 해당 영역이 없으면 null
- total: 총점(정수). 없으면 null
- grade: 등급 표기 그대로 (예: OPIc의 AL/IH/IM3, TOEIC Speaking 레벨 등). 없으면 null

여러 시험이 한 장에 있으면 가장 주된(대표) 결과 1건만 추출하세요."""


class LanguageRecord(BaseModel):
    employee_name: Optional[str] = None
    employee_id: Optional[str] = None
    language: Optional[str] = None
    test_type: Optional[str] = None
    valid_date: Optional[str] = None
    listening: Optional[int] = None
    reading: Optional[int] = None
    writing: Optional[int] = None
    speaking: Optional[int] = None
    total: Optional[int] = None
    grade: Optional[str] = None


def load_env() -> None:
    """프로젝트 폴더의 .env 파일이 있으면 환경변수로 읽어들인다(있을 때만)."""
    env_path = ROOT / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_roster() -> dict:
    """성명->사번 매핑. roster.csv(헤더: 성명,사번)가 있으면 읽는다. 없으면 빈 dict."""
    for candidate in (ROOT / "roster.csv", INPUT_DIR / "roster.csv"):
        if candidate.exists():
            mapping = {}
            with candidate.open(encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    name = (row.get("성명") or row.get("이름") or "").strip()
                    sabun = (row.get("사번") or "").strip()
                    if name:
                        mapping[name] = sabun
            return mapping
    return {}


def extract_record(client, file_path: Path) -> LanguageRecord:
    """한 파일에서 외국어 성적 1건을 추출한다."""
    data = base64.standard_b64encode(file_path.read_bytes()).decode("utf-8")
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        media_block = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": data},
        }
    else:
        media_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": IMAGE_TYPES[ext], "data": data},
        }

    response = client.messages.parse(
        model="claude-opus-4-7",
        max_tokens=6000,
        thinking={"type": "adaptive"},
        system=[{"type": "text", "text": SYSTEM_PROMPT, "cache_control": {"type": "ephemeral"}}],
        messages=[{
            "role": "user",
            "content": [media_block, {"type": "text", "text": "이 외국어 성적표에서 정보를 추출하세요."}],
        }],
        output_format=LanguageRecord,
    )
    record = response.parsed_output
    if record is None:
        raise RuntimeError(f"추출 실패 (stop_reason={response.stop_reason})")
    return record


def resolve_ability(record: LanguageRecord) -> str:
    """외국어 능력(상/중/하)을 등급 -> 점수 순서로 결정. 못 정하면 빈 문자열."""
    if record.grade:
        key = record.grade.upper().replace(" ", "")
        if key in config.ABILITY_BY_GRADE:
            return config.ABILITY_BY_GRADE[key]
    if record.test_type and record.total is not None:
        rules = config.ABILITY_BY_SCORE.get(record.test_type)
        if rules:
            for threshold, ability in rules:
                if record.total >= threshold:
                    return ability
    return ""


def validate(record: LanguageRecord, sabun: str, ability: str, seen: set) -> str:
    """행 단위 검증 메모를 만든다(비어 있으면 문제 없음)."""
    notes = []
    if not record.employee_name:
        notes.append("이름 없음")
    if not sabun:
        notes.append("사번 없음(확인 필요)")
    if not record.test_type:
        notes.append("시험종류 없음")
    if record.total is None and not record.grade:
        notes.append("점수/등급 없음")
    if not ability:
        notes.append("능력(상/중/하) 미결정-규칙 확인")

    if record.valid_date:
        try:
            valid = datetime.date.fromisoformat(record.valid_date)
            if valid < datetime.date.today():
                notes.append(f"만료됨({record.valid_date})")
        except ValueError:
            notes.append("유효일자 형식 확인")
    else:
        notes.append("유효일자 없음")

    key = (record.employee_name, record.test_type)
    if key in seen:
        notes.append("중복 가능")
    seen.add(key)
    return "; ".join(notes)


def to_row(record: LanguageRecord, sabun: str, ability: str, notes: str) -> dict:
    blank = lambda v: "" if v is None else v
    return {
        "사번": sabun or "",
        "성명": record.employee_name or "",
        "외국어 구분": record.language or "",
        "외국어 시험": record.test_type or "",
        "유효일자": record.valid_date or "",
        "듣기점수": blank(record.listening),
        "읽기점수": blank(record.reading),
        "쓰기점수": blank(record.writing),
        "말하기점수": blank(record.speaking),
        "총점": blank(record.total),
        "외국어 등급": record.grade or "",
        "외국어 능력": ability,
        "검증결과": notes,
    }


def write_xlsx(rows: list, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "외국어"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(config.COLUMNS)
    for idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in rows:
        ws.append([row.get(col, "") for col in config.COLUMNS])
        if row.get("검증결과"):  # 문제 있는 행은 연한 주황색으로 표시
            for cell in ws[ws.max_row]:
                cell.fill = warn_fill

    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def build_rows(records: list, roster: dict) -> list:
    rows, seen = [], set()
    for record in records:
        sabun = record.employee_id or roster.get(record.employee_name or "", "")
        ability = resolve_ability(record)
        notes = validate(record, sabun, ability, seen)
        rows.append(to_row(record, sabun, ability, notes))
    return rows


def run(input_dir: Path, out_path: Path) -> None:
    files = sorted(
        p for p in input_dir.iterdir()
        if p.is_file() and p.suffix.lower() in (set(IMAGE_TYPES) | {".pdf"})
    )
    if not files:
        print(f"[안내] '{input_dir}' 폴더에 성적표 이미지(.jpg/.png) 또는 .pdf 를 넣고 다시 실행하세요.")
        return

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("[오류] ANTHROPIC_API_KEY 가 없습니다. .env 파일을 만들거나 환경변수를 설정하세요.")
        print("       (예) .env 파일에  ANTHROPIC_API_KEY=sk-ant-...  한 줄 추가")
        sys.exit(1)

    import anthropic  # API 호출 때만 필요 (--selftest 는 이거 없이도 동작)
    client = anthropic.Anthropic(api_key=api_key)

    roster = load_roster()
    records = []
    for file_path in files:
        print(f"  - 읽는 중: {file_path.name} ...", end=" ", flush=True)
        try:
            records.append(extract_record(client, file_path))
            print("완료")
        except Exception as exc:  # 한 파일이 실패해도 나머지는 계속
            print(f"실패 ({exc})")

    if not records:
        print("[안내] 추출된 결과가 없습니다.")
        return

    rows = build_rows(records, roster)
    write_xlsx(rows, out_path)

    flagged = sum(1 for r in rows if r["검증결과"])
    print(f"\n생성 완료: {out_path}")
    print(f"  총 {len(rows)}건 / 검토 필요 {flagged}건 (주황색 표시 + '검증결과' 칸 확인)")
    if roster:
        print(f"  roster.csv 로 사번 매칭 사용 ({len(roster)}명 등록)")
    else:
        print("  (참고) roster.csv(성명,사번)를 넣으면 사번을 자동으로 채웁니다.")


def selftest(out_path: Path) -> None:
    """API 없이 엑셀 생성/검증 로직만 점검."""
    samples = [
        LanguageRecord(employee_name="이재용", language="영어", test_type="OPIc",
                       valid_date="2026-03-13", speaking=49, total=49, grade="AL"),
        LanguageRecord(employee_name="홍길동", language="영어", test_type="TOEIC",
                       valid_date="2024-01-01", listening=480, reading=470, total=950),
        LanguageRecord(employee_name="김미정", language="영어", test_type="TOEIC Speaking",
                       valid_date=None, total=120),  # 유효일자 없음 + 능력 '하'
    ]
    roster = {"이재용": "82211489", "홍길동": "10001234"}  # 김미정은 누락 -> 사번 없음 경고
    write_xlsx(build_rows(samples, roster), out_path)
    print(f"[selftest] 샘플 3건으로 생성 완료: {out_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="외국어 성적표 -> ERP 업로드 엑셀 변환기")
    parser.add_argument("--input", default=str(INPUT_DIR), help="성적표 폴더 (기본: ./input)")
    parser.add_argument("--output", default=str(OUTPUT_DIR / "외국어_업로드.xlsx"),
                        help="결과 엑셀 경로 (기본: ./output/외국어_업로드.xlsx)")
    parser.add_argument("--selftest", action="store_true", help="API 없이 샘플로 동작 확인")
    args = parser.parse_args()

    load_env()
    out_path = Path(args.output)
    if args.selftest:
        selftest(out_path)
    else:
        run(Path(args.input), out_path)


if __name__ == "__main__":
    main()
